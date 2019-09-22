# -- coding: utf-8 --

# 这个demo练习的是openpyxl包

#读取excle数据并进行打印

from openpyxl  import load_workbook
#打开文件 默认可读写  如果有需要可以指定write_only 和 read_only 为True
wb=load_workbook(filename="excel_operator_demo.xlsx")
#找到工作表
sheet1=wb["Sheet1"]
#打印指定单元格的内容

#print(sheet1["B2"].value)
#按行打印数据    因为按行,  所以返回A1, B1, C1 这样的顺序

# for row in sheet1.rows:
#     for cell in row:
#         print(cell.value)

# #按列打印数据  A1 A2 A3顺序
# for column in sheet1.columns:
#     for  cell in column:
#         print(cell.value)

#只读取某行数据
# for cell in list(sheet1.rows)[2]:
#     print(cell.value)

#只读取某列的数据
# for cell in list(sheet1.columns)[1]:
#     print(cell.value)

#读取任意区间的单元格
# 可以使用range函数，下面的写法，获得了以A1为左上角，B3为右下角矩形区域的所有单元格。
# 注意range从1开始的，因为在openpyxl中为了和Excel中的表达方式一致，并不和编程语言的习惯以0表示第一个值。
# for cell in range(1,4):
#     for i in range(1,3):
#         print(sheet1.cell(row=cell,column=i).value)

# 还可以像使用切片那样使用。
# sheet['A1':'B3']返回一个tuple，该元组内部还是元组，由每行的单元格构成一个元组
#注意测试未成功  目前还没找到相关文档

# for row_cell in sheet1["A1":"B3"]:
#     for cell in row_cell:
#         print(cell)

